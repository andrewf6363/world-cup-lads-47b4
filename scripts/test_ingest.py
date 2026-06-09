#!/usr/bin/env python3
"""Round-trip test: fill copies of the real template with picks (clean + messy), parse them
back with ingest_picks, and assert the parser recovers the intended outcomes."""
import json, os, sys, tempfile
import openpyxl
sys.path.insert(0, os.path.dirname(__file__))
import ingest_picks as ing

HERE = os.path.dirname(__file__)
DATA = os.path.join(HERE, "..", "data")
TEMPLATE = os.path.join(DATA, "blank-template.xlsx")
fixtures = json.load(open(os.path.join(DATA, "fixtures.json")))["group_stage"]
aliases  = json.load(open(os.path.join(DATA, "teams.json")))["aliases"]

# messy variants to stress the resolver (typed value -> should still resolve to that team)
VARIANT = {
    "South Korea":"korea", "Türkiye":"Turkey", "Bosnia-Herzegovina":"Bosnia",
    "Côte d'Ivoire":"Ivory Coast", "Congo DR":"DR Congo", "USA":"United States",
}

def fill(messy):
    """Return (path, expected{id->outcome|'FLAG'|'BLANK'})."""
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb.worksheets[0]
    expected = {}
    for i, fx in enumerate(fixtures):
        r, c = fx["pick_cells"][0]          # type winner in the first pink cell
        outcome = ["team1", "team2", "draw"][i % 3]
        if messy and i == 5:                # one garbage entry -> must be flagged
            ws.cell(row=r, column=c).value = "zzz"; expected[fx["id"]] = "FLAG"; continue
        if messy and i == 9:                # one left blank -> unpicked
            expected[fx["id"]] = "BLANK"; continue
        if outcome == "draw":
            ws.cell(row=r, column=c).value = "TIE" if messy else "tie"
        else:
            team = fx["team1"] if outcome == "team1" else fx["team2"]
            val = team
            if messy:
                val = VARIANT.get(team, team.lower())
            ws.cell(row=r, column=c).value = val
        expected[fx["id"]] = outcome
    path = os.path.join(tempfile.gettempdir(), f"Picks - {'Maria' if messy else 'Diego'}.xlsx")
    wb.save(path)
    return path, expected

def check(label, path, expected):
    name, picks, issues, read = ing.ingest_one(path, fixtures, aliases)
    flagged = {ln.split()[0] for ln in issues}
    errs = []
    for fid, exp in expected.items():
        if exp == "BLANK":
            if fid in picks: errs.append(f"{fid}: blank but parsed {picks[fid]}")
        elif exp == "FLAG":
            if fid not in flagged: errs.append(f"{fid}: garbage not flagged (parsed {picks.get(fid)})")
        else:
            if picks.get(fid) != exp: errs.append(f"{fid}: expected {exp}, got {picks.get(fid)}")
    print(f"[{ 'PASS' if not errs else 'FAIL' }] {label}: {name} · {read}/72 read · {len(issues)} flagged")
    for e in errs[:10]: print("     x", e)
    return not errs

ok = True
ok &= check("clean sheet", *fill(False))
ok &= check("messy sheet (lowercase / partial / alias / TIE / garbage / blank)", *fill(True))
print("\nRESULT:", "ALL PASS" if ok else "FAILURES ABOVE")
sys.exit(0 if ok else 1)
