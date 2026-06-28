#!/usr/bin/env python3
"""Parse a filled knockout bracket spreadsheet into knockout_picks and merge into data/picks.json.

The sheet is a visual bracket: rows 3,4,6,7,... in column B (left half) and V (right half) hold the
32 R32 teams in 16 pairs; advancement is filled rightward — left: D(R32 win) F(R16) H(QF) J(SF),
center L(champion); right: T(R32) R(R16) P(QF) N(SF). We resolve each person's winner for every
bracket match through the fixtures tree by team identity, so a busted pick simply scores 0 downstream.

Usage:  python3 scripts/ingest_knockout.py "/path/to/Knockout Picks - NAME.xlsx" [--name "Full Name"] [--champ "Team"]
"""
import json, os, sys, re, difflib, unicodedata
from openpyxl import load_workbook

HERE = os.path.dirname(__file__); DATA = os.path.join(HERE, "..", "data")
ROWS = [3, 4, 6, 7, 9, 10, 12, 13, 15, 16, 18, 19, 21, 22, 24, 25]   # R32 team rows
# advancement columns (1-indexed): left D=4 F=6 H=8 J=10, champ L=12, right N=14 P=16 R=18 T=20
COL = {"R32": (4, 20), "R16": (6, 18), "QF": (8, 16), "SF": (10, 14)}
CHAMP_COL = 12
ALIASES = {"bosnia": "Bosnia-Herzegovina", "congo": "Congo DR", "drcongo": "Congo DR",
           "ivorycoast": "Côte d'Ivoire", "capeverde": "Cabo Verde", "usa": "USA",
           "unitedstates": "USA", "southkorea": "South Korea", "iran": "IR Iran"}
# the template ships a pre-filled EXAMPLE bracket (England champion); if a person's only filled sheet
# IS this untouched example, they haven't actually made picks — refuse rather than score the default.
DEFAULT_SIG = frozenset((r, c, t) for r, c, t in [
    [3,4,"Germany"],[3,6,"France"],[3,8,"France"],[3,10,"France"],[3,12,"England"],[3,14,"England"],
    [3,16,"England"],[3,18,"Brazil"],[3,20,"Brazil"],[6,4,"France"],[6,20,"Norway"],[9,4,"Canada"],
    [9,6,"Netherlands"],[9,18,"England"],[9,20,"Mexico"],[12,4,"Netherlands"],[12,20,"England"],
    [15,4,"Croatia"],[15,6,"Spain"],[15,8,"Spain"],[15,16,"Argentina"],[15,18,"Argentina"],
    [15,20,"Argentina"],[18,4,"Spain"],[18,20,"Egypt"],[21,4,"USA"],[21,6,"USA"],[21,18,"Colombia"],
    [21,20,"Switzerland"],[24,4,"Belgium"],[24,20,"Colombia"]])

def _sheet_sig(ws):
    cols = [c for pair in COL.values() for c in pair] + [CHAMP_COL]
    return frozenset((r, c, str(ws.cell(row=r, column=c).value).strip())
                     for r in ROWS for c in cols if ws.cell(row=r, column=c).value not in (None, ""))

def norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).strip()
    return "".join(c for c in s.lower() if c.isalnum())

def load(name, default):
    p = os.path.join(DATA, name)
    return json.load(open(p, encoding="utf-8")) if os.path.exists(p) else default

def resolver(valid):
    """Resolve a sheet cell to one of the 32 canonical teams (alias -> exact -> fuzzy-of-32)."""
    by_norm = {norm(t): t for t in valid}
    def res(s):
        if s in (None, ""): return None
        n = norm(s)
        if n in ALIASES and ALIASES[n] in valid: return ALIASES[n]
        if n in by_norm: return by_norm[n]
        m = difflib.get_close_matches(n, list(by_norm), n=1, cutoff=0.7)
        return by_norm[m[0]] if m else ("?" + str(s))      # unresolved -> flagged
    return res

def _adv_filled(ws):
    cols = [c for pair in COL.values() for c in pair] + [CHAMP_COL]
    return sum(1 for r in ROWS for c in cols if ws.cell(row=r, column=c).value not in (None, ""))

def detect_sheet(wb, valid, prefer=None):
    """The bracket the person actually filled. The template ships a pre-filled EXAMPLE sheet ('KB')
    and a blank fill-in sheet ('Master', header L2='Winner'); most people fill Master, a few fill KB
    directly. Rule: an explicit --sheet wins; else a 'Master' sheet with real advancement wins (the
    intended sheet); else the most-filled sheet. Returns (worksheet, note)."""
    cand = [(ws, _adv_filled(ws)) for ws in wb.worksheets]
    if prefer:
        for ws, _ in cand:
            if ws.title.lower() == prefer.lower(): return ws, f"forced sheet '{ws.title}'"
    master = next((ws for ws, adv in cand if ws.title.lower() == "master" and adv >= 8), None)
    if master: return master, f"sheet '{master.title}'"
    ws = max(cand, key=lambda x: x[1])[0]
    note = f"sheet '{ws.title}'" + (" (the blank Master sheet was empty)" if any(c[0].title.lower()=="master" for c in cand) else "")
    return ws, note

def ingest(path, fixtures, name_override=None, champ_override=None, sheet_override=None):
    ko = {k["id"]: k for k in fixtures["knockout"]}
    r32 = [ko[f"K-{n}"] for n in range(73, 89)]
    valid = sorted({t for k in r32 for t in (k["team1"], k["team2"])})
    res = resolver(valid)
    wb = load_workbook(path, data_only=True); ws, sheet_note = detect_sheet(wb, valid, sheet_override)
    if _sheet_sig(ws) == DEFAULT_SIG and not sheet_override:
        return (name_override or re.split(r"\s*-\s*", os.path.splitext(os.path.basename(path))[0])[-1].strip()), \
               {}, None, [], ["NOT SUBMITTED — only the untouched template example is filled (no real bracket)"], "no real bracket"

    # validate the sheet's R32 matchups match the real bracket (by pair)
    sheet_pairs = set()
    col_teams = {2: [], 22: []}
    for col in (2, 22):
        for i in range(0, 16, 2):
            a = res(ws.cell(row=ROWS[i], column=col).value); b = res(ws.cell(row=ROWS[i+1], column=col).value)
            sheet_pairs.add(frozenset((a, b))); col_teams[col] += [a, b]
    real_pairs = {frozenset((k["team1"], k["team2"])) for k in r32}
    bad_pairs = [p for p in sheet_pairs if p not in real_pairs]

    # winners by round (combine both halves into one set per round)
    winners = {}
    flags = []
    for rnd, (lc, rc) in COL.items():
        vals = []
        for col in (lc, rc):
            for r in ROWS:
                v = ws.cell(row=r, column=col).value
                if v not in (None, ""): vals.append(res(v))
        winners[rnd] = [v for v in vals if not str(v).startswith("?")]
        for v in vals:
            if str(v).startswith("?"): flags.append(f"{rnd}: couldn't read '{v[1:]}'")
    champ = champ_override or ws.cell(row=3, column=CHAMP_COL).value
    champ = res(champ) if champ else None
    if not champ or str(champ).startswith("?"):
        flags.append(f"champion not set (cell = {ws.cell(row=3, column=CHAMP_COL).value!r})"); champ = None
    winners["Final"] = [champ] if champ else []

    # resolve each bracket match through the tree by team identity
    picks = {}; cache = {}
    def resolve(mid):
        if mid in cache: return cache[mid]
        k = ko[mid]
        if k["round"] == "R32":
            a, b = k["team1"], k["team2"]
        else:
            a, b = resolve(k["feeds"][0]), resolve(k["feeds"][1])
        wset = set(winners.get(k["round"], []))
        pick = a if a in wset else (b if b in wset else None)
        if pick is None and a and b:
            flags.append(f"{mid} ({k['round']}): pick unresolved for {a} v {b} (bracket inconsistent)")
        cache[mid] = pick
        if pick: picks[mid] = pick
        return pick
    for n in list(range(73, 103)) + [104]:                 # skip 103 (3rd place = 0 pts, not picked)
        resolve(f"K-{n}")

    name = name_override or re.split(r"\s*-\s*", os.path.splitext(os.path.basename(path))[0])[-1].strip()
    return name, picks, champ, bad_pairs, flags, sheet_note

def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    name_ov = next((sys.argv[i+1] for i, a in enumerate(sys.argv) if a == "--name"), None)
    champ_ov = next((sys.argv[i+1] for i, a in enumerate(sys.argv) if a == "--champ"), None)
    sheet_ov = next((sys.argv[i+1] for i, a in enumerate(sys.argv) if a == "--sheet"), None)
    if not args: sys.exit("usage: ingest_knockout.py <file.xlsx> [--name N] [--champ Team] [--sheet KB|Master]")
    fixtures = load("fixtures.json", {"knockout": []})
    if not fixtures.get("knockout"): sys.exit("no knockout fixtures yet — run build_knockout_fixtures.py first")

    name, picks, champ, bad, flags, sheet_note = ingest(args[0], fixtures, name_ov, champ_ov, sheet_ov)
    data = load("picks.json", {"players": []})
    pl = next((p for p in data["players"] if p["name"].lower() == name.lower()), None)
    if not pl: sys.exit(f"'{name}' not found in picks.json (group picks must be ingested first)")
    pl["knockout_picks"] = picks; pl["submitted_knockout"] = bool(picks)   # only "in" with real picks
    json.dump(data, open(os.path.join(DATA, "picks.json"), "w"), ensure_ascii=False, indent=2)

    if not picks:
        print(f"  {name}: NO REAL BRACKET — [{sheet_note}] — left NOT submitted")
        for f in flags: print("    FLAG:", f)
        return
    by_round = {}
    for mid, t in picks.items():
        rnd = next(k["round"] for k in fixtures["knockout"] if k["id"] == mid)
        by_round[rnd] = by_round.get(rnd, 0) + 1
    print(f"  {name}: {len(picks)}/31 picks  (champion: {champ})  [{sheet_note}]  {by_round}")
    if bad: print("    R32 MATCHUP MISMATCH vs real bracket:", bad)
    for f in flags: print("    FLAG:", f)

if __name__ == "__main__":
    main()
